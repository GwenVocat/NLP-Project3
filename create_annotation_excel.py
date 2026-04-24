import pandas as pd
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)

INPUT_TRANSCRIPTIONS = "Data/transcriptions_clean.csv"
INPUT_MAPPING = "Data/ostschweiz_mapping_results.csv"
OUTPUT_EXCEL = "Data/annotation_ground_truth.xlsx"

TOP_N = 100

# ── Colours ──────────────────────────────────────────────────────────────────
FILL_HEADER = PatternFill("solid", fgColor="D9D9D9")
FILL_EDITABLE = PatternFill("solid", fgColor="FFFFC0")
BLUE_BORDER = Side(style="thin", color="4472C4")
BORDER_TOP_BLUE = Border(top=BLUE_BORDER)

WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

COLUMNS = [
    "IPA_Dialekt",
    "Häufigkeit_Total",
    "IPA_Satz",
    "HD_Referenz",
    "Auto_Mapping",
    "Ground_Truth_HD",
    "Korrekt",
    "Notiz",
]

EDITABLE_COLS = {"Ground_Truth_HD", "Korrekt", "Notiz"}

COL_WIDTHS = {
    "IPA_Dialekt":      14,
    "Häufigkeit_Total": 18,
    "IPA_Satz":         52,
    "HD_Referenz":      52,
    "Auto_Mapping":     24,
    "Ground_Truth_HD":  24,
    "Korrekt":          10,
    "Notiz":            28,
}


def load_data():
    trans = pd.read_csv(INPUT_TRANSCRIPTIONS)
    mapping = pd.read_csv(INPUT_MAPPING)
    return trans, mapping


def top_n_ipa_words(mapping_df, n=TOP_N):
    """Return the top-N IPA words by Gemeinsame_Treffer, with their HD mapping."""
    top = (
        mapping_df
        .sort_values("Gemeinsame_Treffer", ascending=False)
        .head(n)
        .reset_index(drop=True)
    )
    return top


def corpus_frequency(trans_df, ipa_words):
    """Count how often each IPA word occurs across all ipa_audio cells."""
    counter = Counter()
    for cell in trans_df["ipa_audio"].dropna():
        tokens = cell.split()
        for tok in tokens:
            if tok in ipa_words:
                counter[tok] += 1
    return counter


def build_rows(trans_df, top_df, freq):
    """Build the flat list of annotation rows (one per occurrence)."""
    # Build a lookup: IPA word → HD auto-mapping
    auto_map = dict(zip(top_df["IPA_Dialekt"], top_df["Hochdeutsch_Zuordnung"]))
    top_words = set(auto_map.keys())

    # Collect occurrences per IPA word
    occurrences: dict[str, list[tuple]] = {w: [] for w in top_words}

    for _, row in trans_df.iterrows():
        ipa_audio = row.get("ipa_audio", "")
        if pd.isna(ipa_audio):
            continue
        tokens = ipa_audio.split()
        seen_in_sentence = set()
        for tok in tokens:
            if tok in top_words and tok not in seen_in_sentence:
                occurrences[tok].append((ipa_audio, row.get("sentence", "")))
                seen_in_sentence.add(tok)

    # Flatten: sort words by corpus frequency descending, then append occurrences
    sorted_words = sorted(top_words, key=lambda w: freq.get(w, 0), reverse=True)

    rows = []
    for word in sorted_words:
        for ipa_satz, hd_ref in occurrences[word]:
            rows.append({
                "IPA_Dialekt":      word,
                "Häufigkeit_Total": freq.get(word, 0),
                "IPA_Satz":         ipa_satz,
                "HD_Referenz":      hd_ref,
                "Auto_Mapping":     auto_map[word],
                "Ground_Truth_HD":  "",
                "Korrekt":          "",
                "Notiz":            "",
            })
    return rows


def write_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Annotation"

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = FILL_HEADER
        cell.alignment = WRAP_TOP_CENTER

    # ── Data rows ─────────────────────────────────────────────────────────────
    current_word = None
    for row_idx, data in enumerate(rows, start=2):
        new_word = data["IPA_Dialekt"] != current_word
        current_word = data["IPA_Dialekt"]

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=data[col_name])

            # Editable columns: yellow background
            if col_name in EDITABLE_COLS:
                cell.fill = FILL_EDITABLE

            # Alignment
            if col_name in ("IPA_Satz", "HD_Referenz", "Ground_Truth_HD", "Notiz"):
                cell.alignment = WRAP_TOP
            else:
                cell.alignment = WRAP_TOP_CENTER

            # Blue top border for first row of each new IPA-word group
            if new_word:
                cell.border = BORDER_TOP_BLUE

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = COL_WIDTHS[col_name]

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_EXCEL)
    print(f"Saved: {OUTPUT_EXCEL}  ({len(rows)} rows)")


def main():
    print("Loading data …")
    trans_df, mapping_df = load_data()

    print(f"Selecting top {TOP_N} IPA words …")
    top_df = top_n_ipa_words(mapping_df, TOP_N)
    top_words = set(top_df["IPA_Dialekt"])

    print("Counting corpus frequencies …")
    freq = corpus_frequency(trans_df, top_words)

    print("Building annotation rows …")
    rows = build_rows(trans_df, top_df, freq)
    print(f"  → {len(rows)} occurrence rows across {len(top_words)} IPA words")

    print("Writing Excel …")
    write_excel(rows)


if __name__ == "__main__":
    main()
